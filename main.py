import sys
import os

if getattr(sys, 'frozen', False):
    # PyInstaller 打包后的资源根目录
    base_path = sys._MEIPASS if hasattr(sys, '_MEIPASS') else os.path.dirname(sys.executable)

    # ---------- 核心修复：纠正 sys.prefix / sys.exec_prefix ----------
    # Paddle 内部会使用这些路径构建库搜索路径，必须设为有效值
    sys.prefix = base_path
    sys.exec_prefix = base_path

    # ----- macOS 动态库搜索路径 -----
    if sys.platform == 'darwin':
        lib_dir = os.path.dirname(sys.executable)
        current = os.environ.get('DYLD_FALLBACK_LIBRARY_PATH', '')
        os.environ['DYLD_FALLBACK_LIBRARY_PATH'] = lib_dir + (os.pathsep + current if current else '')
        print(f"📚 动态库基础路径: {lib_dir}")

    # ----- 模型路径设置 -----
    paddleocr_home = os.path.join(base_path, '.paddleocr')
    if os.path.exists(paddleocr_home):
        os.environ['PADDLEOCR_HOME'] = paddleocr_home
        print(f"PaddleOCR 模型路径: {paddleocr_home}")

    easyocr_home = os.path.join(base_path, '.EasyOCR')
    if os.path.exists(easyocr_home):
        os.environ['EASYOCR_MODULE_PATH'] = easyocr_home
        print(f"EasyOCR 模型路径: {easyocr_home}")

    # ---------- 修复 site.USER_SITE 为 None ----------
    import site
    if site.USER_SITE is None:
        temp_user_site = os.path.join(os.path.dirname(sys.executable), '_user_site')
        os.makedirs(temp_user_site, exist_ok=True)
        site.USER_SITE = temp_user_site
        print(f"📦 临时 USER_SITE: {temp_user_site}")

    # ---------- Paddle 动态库路径显式注入 ----------
    try:
        import paddle
        paddle_root = os.path.dirname(paddle.__file__)
        paddle_libs_dir = os.path.join(paddle_root, 'libs')
        if os.path.exists(paddle_libs_dir):
            # 加入环境变量，确保 Paddle 能找到自己的 .dylib
            os.environ['PADDLE_LIB_PATH'] = paddle_libs_dir
            dyld = os.environ.get('DYLD_FALLBACK_LIBRARY_PATH', '')
            os.environ['DYLD_FALLBACK_LIBRARY_PATH'] = paddle_libs_dir + os.pathsep + dyld
            print(f"📚 Paddle libs 路径: {paddle_libs_dir}")
        else:
            print(f"⚠️ Paddle libs 目录不存在: {paddle_libs_dir}")
    except Exception as e:
        print(f"⚠️ 无法设置 Paddle 动态库路径: {e}")

    # ---------- 环境变量加固（防止 Illegal instruction）----------
    # 这些变量在测试工作流中已设置，但这里再设一次可增加健壮性
    os.environ.setdefault('MKL_DEBUG_CPU_TYPE', '5')
    os.environ.setdefault('MKL_ENABLE_INSTRUCTIONS', 'SSE4_2')
    os.environ.setdefault('OPENBLAS_NUM_THREADS', '1')
    os.environ.setdefault('OMP_NUM_THREADS', '1')
    os.environ.setdefault('KMP_DUPLICATE_LIB_OK', 'TRUE')
    print("🛡️ 兼容性环境变量已就绪")

else:
    os.environ['PADDLEOCR_HOME'] = os.path.expanduser('~/.paddleocr')
    os.environ['EASYOCR_MODULE_PATH'] = os.path.expanduser('~/.EasyOCR')

# macOS 特定配置
if sys.platform == 'darwin':
    os.environ['FLAGS_use_mkldnn'] = '0'
    os.environ['PADDLE_USE_MPS'] = '0'   # x86_64 模式下不启用 MPS
else:
    os.environ['FLAGS_use_mkldnn'] = '0'

os.environ['PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK'] = '1'

import re
import requests
import base64
import uuid
import time
import traceback
import difflib
from pathlib import Path
from typing import Optional, List, Dict, Any
import cv2
import numpy as np
from PIL import Image
import fitz
import pdfplumber

from paddleocr import PaddleOCR
import easyocr

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

import openpyxl
from openpyxl import load_workbook

# ... 后续所有代码保持原样，不需要修改 ...

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# 全局价格映射：商品条码 -> 备案价（数值）
price_map: Dict[str, float] = {}


def load_price_mapping():
    """启动时加载同目录下所有 .xlsx 文件中的商品条码与售价映射"""
    global price_map
    price_map.clear()
    xlsx_files = list(Path('.').glob('*.xlsx'))
    if not xlsx_files:
        print("⚠️ 未找到任何 .xlsx 文件，备案价校验将跳过")
        return

    for file_path in xlsx_files:
        try:
            wb = load_workbook(file_path, data_only=True)
            sheet = wb.active

            # 查找标题行（包含“商品条码”和“售价”的行）
            header_row = None
            # 最多搜索前10行
            max_search_rows = min(10, sheet.max_row)
            for row_idx in range(1, max_search_rows + 1):
                row_contains_barcode = False
                row_contains_price = False
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row_idx, column=col).value
                    if cell_value:
                        cell_str = str(cell_value).strip()
                        if '商品条码' in cell_str:
                            row_contains_barcode = True
                        if '售价' in cell_str or '备案价' in cell_str:
                            row_contains_price = True
                if row_contains_barcode and row_contains_price:
                    header_row = row_idx
                    print(f"✅ 在文件 {file_path.name} 的第 {header_row} 行找到标题行")
                    break

            if header_row is None:
                print(f"⚠️ 文件 {file_path.name} 中未找到包含'商品条码'和'售价'的标题行，跳过")
                continue

            # 确定列映射
            barcode_col = None
            price_col = None
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=header_row, column=col).value
                if cell_value:
                    header = str(cell_value).strip()
                    if '商品条码' in header:
                        barcode_col = col
                    if '售价' in header or '备案价' in header:
                        price_col = col

            if barcode_col is None or price_col is None:
                print(f"⚠️ 文件 {file_path.name} 第{header_row}行未找到必要的列，跳过")
                continue

            # 读取数据行（从标题行的下一行开始）
            for row in range(header_row + 1, sheet.max_row + 1):
                barcode_cell = sheet.cell(row=row, column=barcode_col).value
                price_cell = sheet.cell(row=row, column=price_col).value
                if barcode_cell and price_cell:
                    try:
                        barcode = str(barcode_cell).strip()
                        price = float(price_cell)
                        price_map[barcode] = price
                    except ValueError:
                        print(f"⚠️ 文件 {file_path.name} 第{row}行价格无法转换为数字，跳过")
            print(f"✅ 已加载 {len([k for k in price_map if price_map.get(k)])} 条备案价记录（来自 {file_path.name}）")
        except Exception as e:
            print(f"❌ 读取文件 {file_path.name} 失败: {e}")


# 应用启动时加载价格映射
@app.on_event("startup")
async def startup_event():
    load_price_mapping()
    print("=" * 60)
    print("订单审核助手启动中...")
    print("首次运行会自动下载 OCR 模型（约 100MB），请保持网络畅通")
    print("模型下载需要 2-5 分钟，后续启动将直接使用本地模型")
    print("=" * 60)


# =============================================================================
# 初始化 OCR 引擎
try:
    if sys.platform == 'darwin':
        try:
            # 1. 使用推荐的 OCR 版本
            paddle_ocr = PaddleOCR(lang='ch', ocr_version='PP-OCRv4', show_log=False)
        except:
            paddle_ocr = PaddleOCR(lang='ch')
    else:
        # 2. 弃用 use_textline_orientation 参数
        try:
            paddle_ocr = PaddleOCR(lang='ch', ocr_version='PP-OCRv4', show_log=False)
        except TypeError:
            paddle_ocr = PaddleOCR(lang='ch', show_log=False)
except Exception as e:
    print(f"⚠️ PaddleOCR 初始化失败: {e}")
    paddle_ocr = None


# =============================================================================
# 请求模型
# =============================================================================
class VerifyRequest(BaseModel):
    invoiceUrl: str
    merchantOrderNo: str
    platformOrderNo: str
    sellerName: str
    amount: str
    receiverAddress: str
    productBarcode: str
    buyerName: Optional[str] = None
    buyerTaxId: Optional[str] = None
    snCode: Optional[str] = None
    productCategory: Optional[str] = None
    energyGrade: Optional[str] = None
    totalAmount: Optional[str] = None
    discountAmount: Optional[str] = None
    invoiceImageUrl: Optional[str] = None
    signPurchaseImageUrl: Optional[str] = None
    snImageUrl: Optional[str] = None
    invoiceImageBase64: Optional[str] = None
    signPurchaseImageBase64: Optional[str] = None
    snImageBase64: Optional[str] = None
    # 新增物流字段
    platform: Optional[str] = None
    deliveryMethod: Optional[str] = None
    trackingNo: Optional[str] = None
    logisticsImageUrl: Optional[str] = None


# =============================================================================
# 文件处理工具（优化 macOS 兼容性）
# =============================================================================
def detect_file_type(file_path: Path) -> str:
    """根据文件头检测文件类型，返回扩展名（如'.pdf'）或None"""
    with open(file_path, 'rb') as f:
        header = f.read(8)
    if header.startswith(b'%PDF'):
        return '.pdf'
    if header.startswith(b'\xff\xd8'):
        return '.jpg'
    if header.startswith(b'\x89PNG\r\n\x1a\n'):
        return '.png'
    return None


def ensure_file_extension(file_path: Path, content_type: str = "") -> Path:
    """确保文件有正确的扩展名，如果没有则根据文件头或 Content-Type 添加"""
    if file_path.suffix.lower() in ['.pdf', '.jpg', '.jpeg', '.png', '.bmp', '.ofd']:
        return file_path
    detected_ext = detect_file_type(file_path)
    if detected_ext:
        new_path = file_path.with_suffix(detected_ext)
        file_path.rename(new_path)
        return new_path
    if 'image/jpeg' in content_type:
        new_path = file_path.with_suffix('.jpg')
        file_path.rename(new_path)
        return new_path
    elif 'image/png' in content_type:
        new_path = file_path.with_suffix('.png')
        file_path.rename(new_path)
        return new_path
    elif 'application/pdf' in content_type:
        new_path = file_path.with_suffix('.pdf')
        file_path.rename(new_path)
        return new_path
    return file_path


def download_file(url: str, folder: str) -> Path:
    """下载文件，返回确保有正确扩展名的本地路径"""
    resp = requests.get(url, timeout=30, stream=True)
    resp.raise_for_status()
    filename = None
    content_disposition = resp.headers.get('content-disposition')
    if content_disposition:
        fname_match = re.findall('filename="?([^"]+)"?', content_disposition)
        if fname_match:
            filename = fname_match[0]
    if not filename:
        url_path = url.split('/')[-1].split('?')[0]
        filename = url_path if url_path else 'file'
    save_path = Path(f'./{folder}')
    save_path.mkdir(exist_ok=True)
    file_path = save_path / filename
    with open(file_path, 'wb') as f:
        for chunk in resp.iter_content(1024):
            f.write(chunk)
    content_type = resp.headers.get('content-type', '')
    return ensure_file_extension(file_path, content_type)


def process_image_base64(base64_str: str) -> Path:
    """将 Base64 字符串解码为临时图片文件，返回路径（强制 .png）"""
    if ',' in base64_str:
        base64_data = base64_str.split(',')[1]
    else:
        base64_data = base64_str
    image_data = base64.b64decode(base64_data)
    temp_dir = Path('./temp_images')
    temp_dir.mkdir(exist_ok=True)
    temp_path = temp_dir / f"img_{uuid.uuid4().hex}.png"
    temp_path.write_bytes(image_data)
    return temp_path


def safe_unlink(path: Path, max_retries=3):
    """安全删除文件，重试几次"""
    for i in range(max_retries):
        try:
            if path.exists():
                path.unlink()
            return
        except PermissionError:
            if i < max_retries - 1:
                time.sleep(0.2)
        except Exception:
            pass


# =============================================================================
# OCR 引擎封装（优化版：内存处理 + 智能旋转 + macOS 兼容）
# =============================================================================
MIN_IMAGE_WIDTH = 600  # 调低阈值，减少不必要的放大


def recognize_image_np(img: np.ndarray) -> str:
    """
    对内存中的图像执行 OCR，返回识别文本
    直接使用 PaddleOCR 的 ocr 方法传入图像数组，避免写临时文件
    """
    if paddle_ocr is None:
        return ""
    try:
        # 注意：PaddleOCR 的 ocr 方法支持直接传入 numpy 数组，需要设置 cls=True 启用方向分类
        result = paddle_ocr.ocr(img, cls=True)
        texts = []
        if result and len(result) > 0 and result[0]:
            for line in result[0]:
                if line and len(line) >= 2:
                    text = line[1][0] if isinstance(line[1], tuple) else line[1]
                    confidence = line[1][1] if isinstance(line[1], tuple) and len(line[1]) > 1 else 1.0
                    if confidence > 0.3:
                        texts.append(text)
        return "\n".join(texts)
    except Exception as e:
        print(f"内存 OCR 失败: {e}")
        return ""


def ocr_image_paddle(image_path: Path) -> str:
    """
    使用 PaddleOCR 识别图片，优化版：
    - 直接内存处理
    - 先进行一次识别（方向分类已启用），若文本足够长则返回
    - 若文本过短，再尝试旋转 90° 和 270°（常见方向），取最佳结果
    - 确保不丢失原有多角度覆盖能力，同时大幅减少识别次数
    """
    if paddle_ocr is None:
        print("PaddleOCR 不可用，回退到 EasyOCR")
        return ocr_image_easy(image_path)

    try:
        img = cv2.imread(str(image_path))
        if img is None:
            print("OpenCV 无法读取图像，回退到 EasyOCR")
            return ocr_image_easy(image_path)

        h, w = img.shape[:2]
        print(f"原始图像尺寸: {w} x {h}")

        # 若宽度小于阈值，放大图像（保持长宽比）
        if w < MIN_IMAGE_WIDTH:
            scale = MIN_IMAGE_WIDTH / w
            new_w = MIN_IMAGE_WIDTH
            new_h = int(h * scale)
            img = cv2.resize(img, (new_w, new_h), interpolation=cv2.INTER_CUBIC)
            print(f"图像已放大至: {new_w} x {new_h}")

        # 先进行一次识别（方向分类器自动校正）
        text0 = recognize_image_np(img)
        len0 = len(text0.strip())
        print(f"初始方向识别长度: {len0}")

        best_text = text0
        best_len = len0

        # 如果初始文本长度足够（例如大于100），直接返回，不再旋转
        if best_len >= 100:
            print(f"✅ 初始识别已足够，跳过旋转")
            return best_text

        # 否则，尝试旋转 90° 和 270°
        rotations = [
            (cv2.ROTATE_90_CLOCKWISE, "90°"),
            (cv2.ROTATE_90_COUNTERCLOCKWISE, "270°")
        ]

        for rotate_code, angle_name in rotations:
            current_img = cv2.rotate(img, rotate_code)
            text = recognize_image_np(current_img)
            cur_len = len(text.strip())
            print(f"旋转 {angle_name} 识别结果长度: {cur_len}")
            if cur_len > best_len:
                best_text = text
                best_len = cur_len
                # 如果当前长度已足够，提前终止（可选）
                if best_len >= 100:
                    break

        # 若仍无结果，可考虑旋转180°（但方向分类器应已覆盖，保留注释）
        # 如果上述均无效，最终返回最佳结果
        if best_text:
            print(f"✅ 最佳识别结果长度: {best_len} 字符")
            print(f"最佳识别文本预览: {best_text[:200]}")
        return best_text

    except Exception as e:
        print(f"PaddleOCR 识别失败: {e}，回退到 EasyOCR")
        return ocr_image_easy(image_path)


def ocr_image_easy(image_path: Path) -> str:
    """使用 EasyOCR 识别图片（备选）"""
    if easy_reader is None:
        print("EasyOCR 不可用")
        return ""
    try:
        result = easy_reader.readtext(str(image_path), detail=0, paragraph=True)
        full_text = "\n".join(result) if result else ""
        print(f"EasyOCR 识别结果长度: {len(full_text)} 字符")
        if full_text:
            print(f"EasyOCR 识别文本预览: {full_text[:200]}")
        return full_text
    except Exception as e:
        print(f"EasyOCR 识别失败: {e}")
        return ""


def ocr_image_multi_angle(image_path: Path) -> str:
    """
    对单张图片进行多角度 OCR 识别，合并所有方向检测到的文本。
    适用于方向不确定的图片（如含有横向水印、竖向 SN 码、倒置文字）。
    处理角度：原始、顺时针 90°、180°、逆时针 90°（即顺时针 270°）。
    """
    if paddle_ocr is None:
        return ocr_image_easy(image_path)

    try:
        img = cv2.imread(str(image_path))
        if img is None:
            print("⚠️ 无法读取图片，回退到 EasyOCR")
            return ocr_image_easy(image_path)

        h, w = img.shape[:2]
        # 若宽度过小，适当放大
        if w < MIN_IMAGE_WIDTH:
            scale = MIN_IMAGE_WIDTH / w
            new_w = MIN_IMAGE_WIDTH
            new_h = int(h * scale)
            img = cv2.resize(img, (new_w, new_h), interpolation=cv2.INTER_CUBIC)

        all_texts = []

        def collect_texts(image_np: np.ndarray, angle_desc: str = ""):
            """执行 OCR 并收集文本行"""
            try:
                result = paddle_ocr.ocr(image_np, cls=True)
                if result and len(result) > 0 and result[0]:
                    for line in result[0]:
                        if line and len(line) >= 2:
                            text = line[1][0] if isinstance(line[1], tuple) else line[1]
                            conf = line[1][1] if isinstance(line[1], tuple) and len(line[1]) > 1 else 1.0
                            if conf > 0.3 and text.strip():
                                all_texts.append(text.strip())
            except Exception as e:
                print(f"OCR 子过程失败 ({angle_desc}): {e}")

        # 1. 原始方向
        collect_texts(img, "0°")
        # 2. 顺时针 90°
        collect_texts(cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE), "90°")
        # 3. 180°（旋转两次 90°）
        img_180 = cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
        img_180 = cv2.rotate(img_180, cv2.ROTATE_90_CLOCKWISE)
        collect_texts(img_180, "180°")
        # 4. 逆时针 90°（即顺时针 270°）
        collect_texts(cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE), "270°")

        # 去重（保留首次出现的顺序）
        seen = set()
        unique_texts = []
        for t in all_texts:
            if t not in seen:
                seen.add(t)
                unique_texts.append(t)

        merged = "\n".join(unique_texts)
        print(f"✅ 多角度合并识别完成，共收集 {len(unique_texts)} 条文本")
        return merged

    except Exception as e:
        print(f"多角度 OCR 失败: {e}，回退到 EasyOCR")
        return ocr_image_easy(image_path)

    # =============================================================================


# PDF 处理（优化：降低 DPI）
# =============================================================================
def pdf_to_images(pdf_path: Path, dpi=150) -> list:
    """将 PDF 每页转为灰度图像（默认 DPI 150，提高速度）"""
    images = []
    doc = fitz.open(pdf_path)
    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        pix = page.get_pixmap(dpi=dpi, colorspace="gray")
        img = Image.frombytes("L", [pix.width, pix.height], pix.samples)
        images.append(img)
    doc.close()
    return images


def extract_text_from_pdf(path: Path) -> str:
    """提取 PDF 中的文本，失败则转为图片 OCR（已采用低 DPI 和多线程可选，此处保持单线程以保证稳定性）"""
    try:
        doc = fitz.open(path)
        text = ""
        for page in doc:
            text += page.get_text()
        doc.close()
        if len(text.strip()) >= 50:
            print("✅ 使用PyMuPDF提取文本成功")
            return text
    except Exception as e:
        print(f"⚠️ PyMuPDF打开失败: {e}，尝试pdfplumber")
    try:
        with pdfplumber.open(path) as pdf:
            text = ""
            for page in pdf.pages:
                if page_text := page.extract_text():
                    text += page_text + "\n"
        if len(text.strip()) >= 50:
            print("✅ 使用pdfplumber提取文本成功")
            return text
    except Exception as e:
        print(f"⚠️ pdfplumber打开失败: {e}，转为图片OCR")
    print("📄 文本提取不足或失败，转为图片OCR")
    # 使用优化后的低 DPI 图片
    images = pdf_to_images(path, dpi=150)
    ocr_text = ""
    for i, img in enumerate(images):
        temp_img = path.parent / f"temp_page_{i}.png"
        img.save(temp_img)
        ocr_text += ocr_image_paddle(temp_img) + "\n"
        safe_unlink(temp_img)
    return ocr_text


# =============================================================================
# 字段提取函数（优先使用期望值匹配）
# =============================================================================
def extract_fields(text: str, expected: Dict[str, Any]) -> Dict[str, Any]:
    """
    从发票 OCR 文本中提取字段，优先使用期望值匹配
    expected 包含：buyerName, buyerTaxId, totalAmount, discountAmount, amount, platformOrderNo
    """
    fields = {
        'purchaser': None,
        'buyer_tax_id': None,
        'remark_orders': [],
        'remark_discount': None,
        'remark_total': None,
        'remark_paid': None,
        'extracted_barcode': None,
    }
    print("===== 开始提取字段 =====")
    print("===== 原始文本 =====")
    print(text)
    print("====================")

    # ---------- 购买方名称（优先期望值）----------
    if expected.get('buyerName'):
        # 转义特殊字符，将转义后的 '*' 替换为 '.*' 以支持通配符
        escaped = re.escape(expected['buyerName'])
        pattern = escaped.replace('\\*', '.*')
        candidates = re.findall(r'[\u4e00-\u9fa5]{2,4}', text)
        for cand in candidates:
            if re.fullmatch(pattern, cand):
                fields['purchaser'] = cand
                print(f"✅ 期望值匹配购买方名称: {cand}")
                break

    if not fields['purchaser']:
        name_match = re.search(r'(?:购买方|购货单位).*?名\s*称[：:]\s*([\u4e00-\u9fa5·]{2,20})', text, re.DOTALL)
        if name_match:
            candidate = name_match.group(1)
            if not re.search(r'公司|集团|厂|店|部|中心', candidate):
                fields['purchaser'] = candidate
                print(f"✅ 策略2（名称标签）找到名称: {fields['purchaser']}")

    # ---------- 纳税人识别号（优先期望值）----------
    if expected.get('buyerTaxId'):
        escaped = re.escape(expected['buyerTaxId'])
        pattern = escaped.replace('\\*', '.*')
        tax_candidates = re.findall(r'\b[0-9A-Za-z]{15,20}\b', text)
        for cand in tax_candidates:
            if re.fullmatch(pattern, cand):
                fields['buyer_tax_id'] = cand
                print(f"✅ 期望值匹配纳税人识别号: {cand}")
                break

    if not fields['buyer_tax_id']:
        tax_candidates = re.findall(r'\b[0-9A-Za-z]{15,20}\b', text)
        id_card_pattern = re.compile(r'^\d{17}[\dXx]$|^\d{15}$')
        for cand in tax_candidates:
            if id_card_pattern.match(cand):
                fields['buyer_tax_id'] = cand
                print(f"✅ 找到身份证号（购买方税号）: {cand}")
                break
        if not fields['buyer_tax_id'] and tax_candidates:
            fields['buyer_tax_id'] = tax_candidates[0]
            print(f"⚠️ 未找到身份证，使用第一个税号: {fields['buyer_tax_id']}")

    # ---------- 备注区域提取 ----------
    remark_text = ""
    remark_match = re.search(r'备\s*注[：:]\s*(.*?)(?=\n\s*\n|\n\s*开票人|\Z)', text, re.DOTALL)
    if remark_match:
        remark_text = remark_match.group(1).strip()
    else:
        lines = text.split('\n')
        for i, line in enumerate(lines):
            if '交易订单号' in line or '订单号' in line:
                remark_text = '\n'.join(lines[i:]).strip()
                break
    print("📝 备注区域:", remark_text)

    # 辅助函数：去除非字母数字（用于订单号匹配）
    def alnum(s):
        return re.sub(r'[^A-Za-z0-9]', '', s)

    # ---------- 订单号（优先期望值，忽略符号）----------
    remark_orders = []
    if expected.get('platformOrderNo'):
        expected_alnum = alnum(expected['platformOrderNo'])
        remark_alnum = alnum(remark_text)
        if expected_alnum and expected_alnum in remark_alnum:
            remark_orders = [expected['platformOrderNo']]  # 保留原始格式
            print(f"✅ 期望值（忽略符号）匹配订单号: {expected['platformOrderNo']}")

    if not remark_orders:
        remark_no_space = re.sub(r'\s+', '', remark_text)
        order_candidates = re.findall(r'[A-Za-z0-9\-]{15,50}', remark_no_space)
        filtered_orders = []
        for oc in order_candidates:
            if re.search(r'[A-Za-z]', oc):
                filtered_orders.append(oc)
            elif len(oc) >= 20 and len(oc) <= 30 and oc.isdigit():
                filtered_orders.append(oc)
        filtered_orders = sorted(set(filtered_orders), key=len, reverse=True)
        if filtered_orders:
            remark_orders = filtered_orders
            print("📦 备注中的订单号（原有）:", remark_orders)

    fields['remark_orders'] = remark_orders

    # ---------- 金额提取（优化版：优先期望值匹配全文，再全文价税合计，最后备注兜底）----------
    remark_joined = re.sub(r'\s+', ' ', remark_text)
    remark_joined = re.sub(r'(\d+)\s*\.\s*(\d+)', r'\1.\2', remark_joined)

    # 总金额（订单金额）
    remark_total = None

    # 1. 优先期望值匹配（全文）
    expected_total = expected.get('totalAmount')
    if expected_total:
        expected_clean = re.sub(r'\s+', '', expected_total)
        text_clean = re.sub(r'\s+', '', text)
        if expected_clean in text_clean:
            remark_total = expected_total
            print(f"✅ 期望值匹配订单金额（全文）: {expected_total}")

    # 2. 若未匹配，则从全文中提取价税合计金额（多种正则模式）
    if not remark_total:
        total_patterns = [
            r'价税合计[：:]\s*([\d.]+)',
            r'价税合计[^0-9]*([\d.]+)',
            r'合计[（(]?大写[）)]?[：:]\s*[零壹贰叁肆伍陆柒捌玖拾佰仟万亿元]+.*?小写[：:]\s*([\d.]+)',
            r'合\s*计\s*[（(]?小写[）)]?[：:]\s*([\d.]+)',
            r'¥\s*([\d.]+)\s*$',
            r'金额[（(]?小写[）)]?[：:]\s*([\d.]+)',
        ]
        for pattern in total_patterns:
            match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
            if match:
                remark_total = match.group(1).strip()
                print(f"✅ 从全文提取价税合计金额: {remark_total} (匹配模式: {pattern})")
                break

    # 3. 兜底：从备注区域提取价税合计（原逻辑）
    if not remark_total:
        total_match = re.search(r'价税合计[：:]\s*([\d.]+)', remark_joined)
        if not total_match:
            total_match = re.search(r'合计[：:]\s*([\d.]+)', remark_joined)
        if not total_match:
            total_match = re.search(r'价税合计[^0-9]*([\d.]+)', remark_joined)
        if total_match:
            remark_total = total_match.group(1)
            print(f"⚠️ 从备注区域兜底提取价税合计: {remark_total}")

    fields['remark_total'] = remark_total

    # 优惠金额（保持原逻辑，优先期望值匹配备注区域）
    if expected.get('discountAmount'):
        expected_no_space = re.sub(r'\s+', '', expected['discountAmount'])
        if expected_no_space in re.sub(r'\s+', '', remark_text):
            fields['remark_discount'] = expected['discountAmount']
            print(f"✅ 期望值匹配优惠金额: {expected['discountAmount']}")
    if not fields['remark_discount']:
        discount_match = re.search(r'(?:政府补贴|优惠)[：:]\s*([\d.]+)', remark_joined)
        if discount_match:
            fields['remark_discount'] = discount_match.group(1)

    # 实付金额（保持原逻辑，优先期望值匹配备注区域）
    if expected.get('amount'):
        expected_no_space = re.sub(r'\s+', '', expected['amount'])
        if expected_no_space in re.sub(r'\s+', '', remark_text):
            fields['remark_paid'] = expected['amount']
            print(f"✅ 期望值匹配实付金额: {expected['amount']}")
    if not fields['remark_paid']:
        paid_match = re.search(r'(?:顾客实付|实付金额)[：:]\s*([\d.]+)', remark_joined)
        if not paid_match:
            paid_match = re.search(r'实付[：:]\s*([\d.]+)', remark_joined)
        if not paid_match:
            paid_match = re.search(r'实付[^0-9]*([\d.]+)', remark_joined)
        if paid_match:
            fields['remark_paid'] = paid_match.group(1)

    print("💰 金额提取:", {
        'total': fields['remark_total'],
        'discount': fields['remark_discount'],
        'paid': fields['remark_paid']
    })

    # ---------- 商品条码提取 ----------
    barcode_match = re.search(r'69\d{13}', text)
    if barcode_match:
        fields['extracted_barcode'] = barcode_match.group()
    else:
        barcode_match = re.search(r'69(\d{11})', text)
        if barcode_match:
            fields['extracted_barcode'] = '69' + barcode_match.group(1)

    print("🎯 最终提取结果:", fields)
    return fields

# =============================================================================
# 签购单字段提取
# =============================================================================
def extract_sign_fields(ocr_text: str, expected: Dict[str, Any]) -> Dict[str, Optional[str]]:
    """
    从签购单 OCR 文本中提取字段，优先使用期望值匹配，支持模糊匹配处理 OCR 错误。
    expected 包含：platformOrderNo, totalAmount, amount, discountAmount
    """
    result = {
        'order': None,          # 银商订单号
        'external_order': None, # 外部订单号
        'original': None,
        'paid': None,
        'discount': None
    }
    print("签购单 OCR 文本:\n", ocr_text)

    ocr_no_space = re.sub(r'\s+', '', ocr_text)

    # ----- 辅助函数：清理并修复常见 OCR 数字混淆 -----
    def clean_for_matching(s: str) -> str:
        """去除空格、转大写，并修复数字与字母混淆"""
        if not s:
            return s
        s = re.sub(r'\s+', '', s).upper()
        # 常见混淆映射
        confusions = {
            'O': '0',      # 字母 O → 数字 0
            'I': '1',      # 字母 I → 数字 1
            'L': '1',      # 字母 L → 数字 1
            'S': '5',      # 字母 S → 数字 5
            'Z': '2',      # 字母 Z → 数字 2
            'B': '8',      # 字母 B → 数字 8
            'G': '6',      # 字母 G → 数字 6
            'Q': '0',      # 字母 Q → 数字 0
            'D': '0',      # 字母 D → 数字 0
        }
        for k, v in confusions.items():
            s = s.replace(k, v)
        return s

    def fuzzy_match_expected(expected_val: str, source_text: str, threshold=0.85) -> bool:
        """判断期望值是否与文本中某子串模糊匹配"""
        if not expected_val or not source_text:
            return False
        exp_clean = clean_for_matching(expected_val)
        src_clean = clean_for_matching(source_text)
        # 直接子串包含（忽略大小写、混淆后）
        if exp_clean in src_clean:
            return True
        # 滑动窗口计算最大相似度
        best_ratio = 0.0
        exp_len = len(exp_clean)
        src_len = len(src_clean)
        if exp_len <= src_len:
            for i in range(src_len - exp_len + 1):
                window = src_clean[i:i+exp_len]
                ratio = difflib.SequenceMatcher(None, exp_clean, window).ratio()
                if ratio > best_ratio:
                    best_ratio = ratio
        return best_ratio >= threshold

    # ----- 订单号期望值匹配（增强版）-----
    platform_order = expected.get('platformOrderNo')
    if platform_order:
        expected_no_space = re.sub(r'\s+', '', platform_order)
        # 1. 精确忽略空格匹配
        if expected_no_space in ocr_no_space:
            result['order'] = platform_order
            result['external_order'] = platform_order
            print(f"✅ 期望值精确匹配订单号: {platform_order}")
        else:
            # 2. 模糊匹配
            if fuzzy_match_expected(platform_order, ocr_text, threshold=0.85):
                result['order'] = platform_order
                result['external_order'] = platform_order
                print(f"✅ 期望值模糊匹配订单号: {platform_order}")

    # 如果未匹配，再用正则提取银商订单号和外部订单号
    if not result['order']:
        order_match = re.search(r'银商订单号[:：]?\s*(\d{20,30})', ocr_text)
        if not order_match:
            order_match = re.search(r'\b(2026\d{22})\b', ocr_text)
        if order_match:
            result['order'] = order_match.group(1)
    if not result['external_order']:
        ext_match = re.search(r'外部订单号[:：]?\s*(\w{20,40})', ocr_text)
        if ext_match:
            result['external_order'] = ext_match.group(1)
        else:
            ext_match = re.search(r'外部订单号[^0-9A-Za-z]*([0-9A-Za-z]{20,40})', ocr_text)
            if ext_match:
                result['external_order'] = ext_match.group(1)

    # 若仍无结果，尝试提取最长数字串作为备选
    if not result['order'] and not result['external_order']:
        long_numbers = re.findall(r'\b(\d{20,30})\b', ocr_text)
        if long_numbers:
            result['order'] = max(long_numbers, key=len)

    # ----- 金额期望值模糊匹配 -----
    def fuzzy_match_amount(expected_amt: str, source_text: str) -> bool:
        """专门针对金额的模糊匹配，只关注数字部分"""
        if not expected_amt:
            return False
        exp_digits = re.sub(r'[^\d.]', '', expected_amt)
        src_digits = re.sub(r'[^\d.]', '', source_text)
        return fuzzy_match_expected(exp_digits, src_digits, threshold=0.85)

    # 原始金额
    if expected.get('totalAmount'):
        if fuzzy_match_amount(expected['totalAmount'], ocr_text):
            result['original'] = expected['totalAmount']
            print(f"✅ 期望值匹配原始金额: {expected['totalAmount']}")
    if not result['original']:
        orig_match = re.search(r'金额[（(]?RMB[）)]?\s*[:：]?\s*([\d.]+)', ocr_text)
        if not orig_match:
            orig_match = re.search(r'原始金额[^0-9]*([\d.]+)', ocr_text)
        if orig_match:
            result['original'] = orig_match.group(1)

    # 实付金额
    if expected.get('amount'):
        if fuzzy_match_amount(expected['amount'], ocr_text):
            result['paid'] = expected['amount']
            print(f"✅ 期望值匹配实付金额: {expected['amount']}")
    if not result['paid']:
        paid_match = re.search(r'支付[\s]*([\d.]+)', ocr_text)
        if not paid_match:
            paid_match = re.search(r'银联二维码支付[\s]*([\d.]+)', ocr_text)
        if paid_match:
            result['paid'] = paid_match.group(1)
        else:
            numbers = re.findall(r'([\d.]+)', ocr_text)
            if numbers:
                result['paid'] = numbers[-1]

    # 优惠金额
    if expected.get('discountAmount'):
        if fuzzy_match_amount(expected['discountAmount'], ocr_text):
            result['discount'] = expected['discountAmount']
            print(f"✅ 期望值匹配优惠金额: {expected['discountAmount']}")
    if not result['discount']:
        disc_match = re.search(r'优惠[\s]*([\d.]+)', ocr_text)
        if not disc_match:
            disc_match = re.search(r'补贴[\s]*([\d.]+)', ocr_text)
        if disc_match:
            result['discount'] = disc_match.group(1)

    return result

def extract_sn_fields(ocr_text: str, expected_sn: Optional[str] = None) -> Optional[str]:
    """
    从 SN 码图片 OCR 文本中提取 SN 码，优先使用期望值匹配（忽略空格和轻微 OCR 错误）
    """
    print("SN码图片 OCR 文本:\n", ocr_text)

    # 辅助函数：清理字符串，去除空格，转小写
    def clean(s):
        return re.sub(r'\s+', '', s).lower()

    # 1. 期望值匹配（忽略空格）
    if expected_sn:
        expected_clean = clean(expected_sn)
        ocr_clean = clean(ocr_text)
        # 精确子串匹配
        if expected_clean in ocr_clean:
            print(f"✅ 期望值（忽略空格）匹配成功: {expected_sn}")
            return expected_sn

        # 模糊匹配：在 OCR 清理后的字符串中寻找与期望值相似度较高的子串
        best_ratio = 0.0
        best_match = None
        # 滑动窗口，长度与期望值相同
        if len(expected_clean) <= len(ocr_clean):
            for i in range(len(ocr_clean) - len(expected_clean) + 1):
                window = ocr_clean[i:i+len(expected_clean)]
                ratio = difflib.SequenceMatcher(None, expected_clean, window).ratio()
                if ratio > best_ratio:
                    best_ratio = ratio
                    best_match = window
        # 也尝试长度在期望值 0.8~1.2 倍范围内的窗口
        min_len = max(6, int(len(expected_clean) * 0.8))
        max_len = min(len(ocr_clean), int(len(expected_clean) * 1.2))
        for length in range(min_len, max_len+1):
            if length > len(ocr_clean):
                continue
            for i in range(len(ocr_clean) - length + 1):
                window = ocr_clean[i:i+length]
                ratio = difflib.SequenceMatcher(None, expected_clean, window).ratio()
                if ratio > best_ratio:
                    best_ratio = ratio
                    best_match = window

        if best_match and best_ratio > 0.7:
            print(f"✅ 模糊匹配成功 (相似度 {best_ratio:.2f}): 期望 {expected_sn}, 匹配 {best_match}")
            return expected_sn

    # 2. 自动提取：查找关键词行
    sn_patterns = [
        r'(?:Serial\s*Number|S/N|SN)[：:]\s*([A-Za-z0-9\s]+)',
        r'(?:序列号)[：:]\s*([A-Za-z0-9\s]+)'
    ]
    for pattern in sn_patterns:
        match = re.search(pattern, ocr_text, re.IGNORECASE)
        if match:
            candidate = match.group(1).strip()
            candidate_clean = re.sub(r'\s+', '', candidate)
            if candidate_clean:
                print(f"✅ 从关键词行提取 SN: {candidate_clean}")
                return candidate_clean

    # 3. 通用自动提取：所有长度 6-30 位的字母数字组合（词边界）
    all_candidates = re.findall(r'\b[A-Za-z0-9]{6,30}\b', ocr_text)
    if not all_candidates:
        return None

    # 优先选择包含字母的候选
    for cand in all_candidates:
        if re.search(r'[A-Za-z]', cand):
            print(f"✅ 提取包含字母的SN码: {cand}")
            return cand

    # 若全是数字，排除商品条码
    non_barcode = [c for c in all_candidates if not (len(c) == 13 and c.startswith('69') and c.isdigit())]
    if non_barcode:
        sn = max(non_barcode, key=len)
        print(f"⚠️ 提取非69开头的数字串作为SN码: {sn}")
        return sn

    print(f"⚠️ 默认取第一个候选: {all_candidates[0]}")
    return all_candidates[0]

# =============================================================================
# 新增：水印地址提取
# =============================================================================
def extract_watermark_address(ocr_text: str) -> Optional[str]:
    """从 SN 码图片 OCR 文本中提取水印地址（真正的收货地址，而非厂商地址）"""
    lines = ocr_text.split('\n')
    # 排除厂商/公司相关的关键词
    exclude_keywords = ['公司', '厂', '有限公司', '制造商', '生产厂', '厂家']
    # 地址特征关键词
    address_keywords = ['区', '县', '路', '街', '镇', '乡', '村', '大道', '巷', '号', '单元', '栋']
    
    best_line = None
    for line in lines:
        line = line.strip()
        if '重庆' not in line:
            continue
        # 如果包含排除关键词，跳过
        if any(kw in line for kw in exclude_keywords):
            continue
        # 优先选择包含地址特征的行
        if any(kw in line for kw in address_keywords):
            best_line = line
            break
        # 否则，长度大于10的行作为备选
        if len(line) > 10 and best_line is None:
            best_line = line
    
    return best_line

# =============================================================================
# 新增：物流图片字段提取
# =============================================================================
def extract_logistics_fields(ocr_text: str, expected: Dict[str, Any]) -> Dict[str, Any]:
    """
    从物流信息图片 OCR 文本中提取字段，优先使用期望值匹配
    expected 包含：trackingNo, buyerName
    """
    result = {
        'tracking_no': None,
        'receiver_name': None,
        'receiver_address': None,
        'status': None,
    }
    print("物流图片 OCR 文本:\n", ocr_text)

    ocr_no_space = re.sub(r'\s+', '', ocr_text)

    # ---------- 物流单号 ----------
    expected_tracking = expected.get('trackingNo')
    if expected_tracking:
        expected_no_space = re.sub(r'\s+', '', expected_tracking)
        if expected_no_space in ocr_no_space:
            result['tracking_no'] = expected_tracking
            print(f"✅ 期望值匹配物流单号: {expected_tracking}")

    if not result['tracking_no']:
        # 尝试匹配常见的物流单号格式（字母数字组合，长度10-30）
        candidates = re.findall(r'\b([A-Za-z0-9]{10,30})\b', ocr_text)
        # 优先选择包含字母的（如SF开头）
        for cand in candidates:
            if re.search(r'[A-Za-z]', cand):
                result['tracking_no'] = cand
                print(f"✅ 提取包含字母的物流单号: {cand}")
                break
        if not result['tracking_no'] and candidates:
            # 纯数字单号，优先选长度接近常见长度的
            result['tracking_no'] = max(candidates, key=len)
            print(f"⚠️ 提取最长数字串作为物流单号: {result['tracking_no']}")

    # ---------- 收货人 ----------
    expected_buyer = expected.get('buyerName')
    if expected_buyer:
        pattern = re.escape(expected_buyer).replace('\\*', '.*')
        name_candidates = re.findall(r'[\u4e00-\u9fa5]{2,4}', ocr_text)
        for cand in name_candidates:
            if re.fullmatch(pattern, cand):
                result['receiver_name'] = cand
                print(f"✅ 期望值匹配收货人: {cand}")
                break

    if not result['receiver_name']:
        # 多种收货人标签匹配
        match = re.search(r'(?:收货人|收件人|联系人)[：:]\s*([\u4e00-\u9fa5·]{2,10})', ocr_text)
        if match:
            result['receiver_name'] = match.group(1)
            print(f"✅ 关键词提取收货人: {result['receiver_name']}")
        else:
            # 尝试从 "收货地址：张三 138xxxx 重庆..." 这种格式中提取
            addr_line_match = re.search(r'收货地址[：:]\s*([\u4e00-\u9fa5·]{2,4})\s*[\d\s]', ocr_text)
            if addr_line_match:
                result['receiver_name'] = addr_line_match.group(1)
                print(f"✅ 从收货地址行提取收货人: {result['receiver_name']}")

    # ---------- 收货地址 ----------
    addr_match = re.search(r'详细地址[：:]\s*([^\n]+)', ocr_text)
    if not addr_match:
        addr_match = re.search(r'地址[：:]\s*([^\n]+)', ocr_text)
    if not addr_match:
        # 若没有明确标签，则取包含"重庆"且长度大于10的行
        lines_temp = ocr_text.split('\n')
        for line in lines_temp:
            if '重庆' in line and len(line.strip()) > 10:
                addr_match = re.search(r'(.*重庆.*)', line)
                break
    if addr_match:
        # 清理可能附带的电话号码等
        raw_addr = addr_match.group(1).strip()
        # 移除电话号码（连续11位数字）
        raw_addr = re.sub(r'\b1\d{10}\b', '', raw_addr).strip()
        result['receiver_address'] = raw_addr
        print(f"✅ 提取收货地址: {result['receiver_address']}")
    else:
        # 最后兜底：取最长行（长度>10且含中文）
        lines_temp = [l.strip() for l in ocr_text.split('\n') if l.strip()]
        long_lines = [l for l in lines_temp if len(l) > 10 and re.search(r'[\u4e00-\u9fa5]', l)]
        if long_lines:
            result['receiver_address'] = max(long_lines, key=len)
            print(f"⚠️ 默认取最长行作为地址: {result['receiver_address']}")

    # ---------- 状态提取（增强时间戳解析 + 状态有效性过滤） ----------
    finish_keywords = [
        '已签收', '签收', '已妥投', '妥投', '已送达', '送达',
        '已派送', '派送完成', '派件完成', '快件已签收',
        '已收货', '收货', '本人签收', '家人签收', '已取件',
        '已完成', '完结', '已投递', '投递', '已签收完毕'
    ]
    abnormal_keywords = ['退回', '拒收', '异常', '未签收', '派送失败', '无法派送', '滞留']
    exclude_keywords = ['预计', '预约', '再派送', '再次派送', '延迟', '转寄', '配送时间更改为']

    # 状态有效性判定：必须包含中文且包含物流相关动作词，且不是纯地址/单号
    def is_valid_status(text: str) -> bool:
        if not text:
            return False
        text = text.strip()
        if len(text) < 2:
            return False
        # 必须包含中文字符
        if not re.search(r'[\u4e00-\u9fa5]', text):
            return False
        # 必须包含物流动作关键词（如签收、派送、运输、揽件等）
        action_words = finish_keywords + abnormal_keywords + ['运输', '转运', '发出', '揽收', '揽件', '出库', '入库', '处理', '接单', '下单']
        if not any(kw in text for kw in action_words):
            return False
        # 排除明显的地址行（若包含"重庆"且同时含有"区/县/路/街/号"等，但注意有些状态可能也含地址，如"【重庆】已签收"，故不能一刀切排除，此处仅排除纯地址无动作词的行，上面动作词检查已覆盖）
        return True

    lines = ocr_text.splitlines()
    timeline = []  # (timestamp_str, status_text)

    # 增强的时间戳正则：支持多种紧凑格式
    timestamp_pattern = re.compile(
        r'(?:时间[：:]\s*)?'                           # 可选的时间前缀
        r'(\d{4}[-/年]\d{1,2}[-/月]\d{1,2}[日]?)'      # 日期部分
        r'\s*'                                          # 可能无空格
        r'(\d{1,2}[:：]\d{1,2}(?:[:：]\d{1,2})?)'      # 时间部分
        r'\s*'                                          # 可能空格
        r'(.*)'                                         # 状态文本
    )

    for i, line in enumerate(lines):
        line_stripped = line.strip()
        if not line_stripped:
            continue

        m = timestamp_pattern.search(line_stripped)
        if m:
            date_part = m.group(1).replace('年', '-').replace('月', '-').replace('日', '')
            time_part = m.group(2).replace('：', ':')
            timestamp = f"{date_part} {time_part}"
            status_text = m.group(3).strip()

            # 若状态文本为空，尝试合并下一行
            if not status_text and i + 1 < len(lines):
                next_line = lines[i+1].strip()
                # 下一行不应包含新的时间戳
                if not timestamp_pattern.search(next_line):
                    status_text = next_line

            # 若合并后仍然无文本，则取当前行剩余部分作为状态
            if not status_text:
                # 例如 "已签收2026-04-16 16:56:58" 这种情况，状态可能在时间之前
                before_ts = line_stripped[:m.start()].strip()
                if before_ts:
                    status_text = before_ts

            # **关键过滤：只有状态文本有效时才加入时间线**
            if is_valid_status(status_text):
                timeline.append((timestamp, status_text))
            else:
                print(f"⚠️ 过滤无效状态候选: {status_text} (行: {line_stripped})")
            continue

        # 有些日志中时间戳和状态文本可能被OCR拆成两行，例如：
        # "已签收"
        # "2026-04-16 16:56:58"
        pure_ts_match = re.match(r'^(\d{4}-\d{1,2}-\d{1,2})\s+(\d{1,2}:\d{1,2}(?::\d{1,2})?)$', line_stripped)
        if pure_ts_match and i > 0:
            prev_line = lines[i-1].strip()
            if any(kw in prev_line for kw in finish_keywords + abnormal_keywords):
                timestamp = f"{pure_ts_match.group(1)} {pure_ts_match.group(2)}"
                if is_valid_status(prev_line):
                    timeline.append((timestamp, prev_line))
                    print(f"🔧 修复跨行时间戳: {timestamp} -> {prev_line}")

    # 若 timeline 不为空，按时间排序取最新状态
    if timeline:
        timeline.sort(key=lambda x: x[0])
        # 从后往前找到第一个不包含排除关键词的状态
        latest_status = None
        for ts, status in reversed(timeline):
            if not any(ek in status for ek in exclude_keywords):
                latest_status = status
                break
        if latest_status is None:
            latest_status = timeline[-1][1]
        result['status'] = latest_status
        print(f"✅ 基于时间排序提取最新状态: {latest_status} (共{len(timeline)}条记录)")
    else:
        # 无时间戳时，按关键词优先级搜索（不依赖行顺序）
        print("⚠️ 未提取到时间戳，按关键词优先级搜索全文")
        found_status = None

        # 优先查找完结状态关键词
        for kw in finish_keywords:
            if kw in ocr_text:
                # 提取包含该关键词的整行或上下文
                match_line = re.search(rf'^.*{re.escape(kw)}.*$', ocr_text, re.MULTILINE)
                if match_line:
                    candidate = match_line.group(0).strip()
                    if is_valid_status(candidate):
                        found_status = candidate
                        print(f"✅ 关键词匹配完结状态: {found_status}")
                        break

        if not found_status:
            # 次选异常状态
            for kw in abnormal_keywords:
                if kw in ocr_text:
                    match_line = re.search(rf'^.*{re.escape(kw)}.*$', ocr_text, re.MULTILINE)
                    if match_line:
                        candidate = match_line.group(0).strip()
                        if is_valid_status(candidate):
                            found_status = candidate
                            print(f"⚠️ 关键词匹配异常状态: {found_status}")
                            break

        if not found_status:
            # 最后取文本中第一行非空且非地址行
            for line in lines:
                line_clean = line.strip()
                if not line_clean or '收货地址' in line_clean:
                    continue
                if any(kw in line_clean for kw in exclude_keywords):
                    continue
                if is_valid_status(line_clean):
                    found_status = line_clean
                    print(f"⚠️ 取首行有效状态行: {found_status}")
                    break

        result['status'] = found_status or ''

    return result
# =============================================================================
# 验证辅助函数
# =============================================================================
def validate_purchaser_name(name: str) -> bool:
    if not name: return False
    if re.search(r'公司|集团|厂|店|部|中心', name): return False
    return bool(re.fullmatch(r'[\u4e00-\u9fa5·]{2,10}', name))

def validate_id_number(id_num: str) -> bool:
    if not id_num: return False
    return bool(re.fullmatch(r'\d{17}[\dXx]', id_num))

def match_with_mask(extracted: str, expected: str) -> bool:
    if not expected:
        return False
    if '*' not in expected:
        return extracted == expected
    pattern = '^' + re.escape(expected).replace('\\*', '.') + '$'
    return bool(re.match(pattern, extracted))

def clean_number_str(s: Optional[str]) -> Optional[str]:
    """清理金额字符串，只保留数字和小数点，移除末尾无效字符"""
    if not s:
        return None
    cleaned = re.sub(r'[^\d.-]', '', s)
    if cleaned.endswith('.'):
        cleaned = cleaned[:-1]
    parts = cleaned.split('.')
    if len(parts) > 2:
        cleaned = parts[0] + '.' + ''.join(parts[1:])
    return cleaned

def make_json_serializable(obj):
    if obj is None or isinstance(obj, (str, int, float, bool)):
        return obj
    if isinstance(obj, (list, tuple)):
        return [make_json_serializable(item) for item in obj]
    if isinstance(obj, dict):
        return {k: make_json_serializable(v) for k, v in obj.items()}
    return str(obj)

# =============================================================================
# 主验证接口
# =============================================================================
@app.post("/verify")
async def verify_invoice(req: VerifyRequest):
    try:
        # ---------- 发票处理 ----------
        invoice_text = ""
        extracted = {}
        if req.invoiceImageBase64:
            try:
                img_path = process_image_base64(req.invoiceImageBase64)
                invoice_text = ocr_image_paddle(img_path)
                safe_unlink(img_path)
                print("✅ 发票 Base64 处理成功")
            except Exception as e:
                print(f"发票 Base64 处理失败: {e}")

        if not invoice_text and req.invoiceUrl:
            try:
                file_path = download_file(req.invoiceUrl, 'temp_invoices')
                ext = file_path.suffix.lower()
                print(f"发票文件类型: {ext}")
                if ext == '.pdf':
                    invoice_text = extract_text_from_pdf(file_path)
                elif ext in ['.jpg', '.jpeg', '.png', '.bmp']:
                    invoice_text = ocr_image_paddle(file_path)
                else:
                    invoice_text = extract_text_from_pdf(file_path)
                    if not invoice_text:
                        invoice_text = ocr_image_paddle(file_path)
                safe_unlink(file_path)
            except Exception as e:
                print(f"发票 URL 处理失败: {e}")

        if invoice_text:
            expected = {
                'buyerName': req.buyerName,
                'buyerTaxId': req.buyerTaxId,
                'totalAmount': req.totalAmount,
                'discountAmount': req.discountAmount,
                'amount': req.amount,
                'platformOrderNo': req.platformOrderNo,
                'snCode': req.snCode,
            }
            extracted = extract_fields(invoice_text, expected)
            print("发票提取字段:", extracted)
        else:
            extracted = {}

        # ---------- 签购单处理 ----------
        sign_order = None
        sign_external = None
        sign_original = None
        sign_paid = None
        sign_discount = None

        if req.signPurchaseImageBase64 or req.signPurchaseImageUrl:
            try:
                if req.signPurchaseImageBase64:
                    img_path = process_image_base64(req.signPurchaseImageBase64)
                else:
                    img_path = download_file(req.signPurchaseImageUrl, 'temp_images')
                sign_ocr = ocr_image_paddle(img_path)
                safe_unlink(img_path)
                expected_sign = {
                    'platformOrderNo': req.platformOrderNo,
                    'totalAmount': req.totalAmount,
                    'amount': req.amount,
                    'discountAmount': req.discountAmount,
                }
                sign_fields = extract_sign_fields(sign_ocr, expected_sign)
                sign_order = sign_fields['order']
                sign_external = sign_fields['external_order']
                sign_original = sign_fields['original']
                sign_paid = sign_fields['paid']
                sign_discount = sign_fields['discount']
                print(f"签购单提取: 银商订单号={sign_order}, 外部订单号={sign_external}, 原始={sign_original}, 实付={sign_paid}, 优惠={sign_discount}")
            except Exception as e:
                print(f"签购单处理失败: {e}")

        # ---------- SN 码图片处理 ----------
        sn_from_image = None
        sn_watermark_address = None
        if req.snImageBase64 or req.snImageUrl:
            try:
                if req.snImageBase64:
                    img_path = process_image_base64(req.snImageBase64)
                else:
                    img_path = download_file(req.snImageUrl, 'temp_images')
                
                # 第一步：尝试原始方向 OCR
                sn_ocr = ocr_image_paddle(img_path)
                sn_from_image = extract_sn_fields(sn_ocr, req.snCode)
                sn_watermark_address = extract_watermark_address(sn_ocr)
                
                # 判断是否已同时满足：SN码匹配 + 地址包含重庆
                sn_ok = sn_from_image is not None and sn_from_image == req.snCode
                addr_ok = sn_watermark_address is not None and '重庆' in sn_watermark_address
                
                if not (sn_ok and addr_ok):
                    print("⚠️ 原始方向未同时提取到 SN 码和重庆地址，启动多角度 OCR...")
                    # 执行多角度合并识别
                    sn_ocr = ocr_image_multi_angle(img_path)
                    sn_from_image = extract_sn_fields(sn_ocr, req.snCode)
                    sn_watermark_address = extract_watermark_address(sn_ocr)
                else:
                    print("✅ 原始方向已同时提取到 SN 码和重庆地址，跳过旋转")
                
                safe_unlink(img_path)
                print(f"SN码提取: {sn_from_image}")
                print(f"水印地址提取: {sn_watermark_address}")
            except Exception as e:
                print(f"SN码图片处理失败: {e}")

        # ---------- 物流信息审核（新增） ----------
        # 仅当线上平台且提货方式非“自提”且存在物流图片 URL 时执行
        logistics_fields = {}
        if req.platform == 'online' and req.deliveryMethod != '自提' and req.logisticsImageUrl:
            try:
                file_path = download_file(req.logisticsImageUrl, 'temp_logistics')
                logistics_ocr = ocr_image_paddle(file_path)
                safe_unlink(file_path)

                expected_logistics = {
                    'trackingNo': req.trackingNo,
                    'buyerName': req.buyerName,
                }
                logistics_fields = extract_logistics_fields(logistics_ocr, expected_logistics)or {}
                print("物流提取字段:", logistics_fields)
            except Exception as e:
                print(f"物流图片处理失败: {e}")
                # 处理失败时标记为不通过，但继续后续流程

        # ---------- 验证详情 ----------
        details = {}

        # 购买方名称
        purchaser_passed = False
        purchaser_extracted = extracted.get('purchaser')
        if req.buyerName:
            purchaser_passed = match_with_mask(purchaser_extracted or '', req.buyerName)
        else:
            purchaser_passed = validate_purchaser_name(purchaser_extracted)
        details['purchaser'] = {'passed': purchaser_passed, 'extracted': purchaser_extracted, 'expected': req.buyerName}

        # 纳税人识别号
        tax_id_passed = False
        tax_id_extracted = extracted.get('buyer_tax_id')
        if req.buyerTaxId:
            tax_id_passed = match_with_mask(tax_id_extracted or '', req.buyerTaxId)
        else:
            tax_id_passed = validate_id_number(tax_id_extracted)
        details['buyer_tax_id'] = {'passed': tax_id_passed, 'extracted': tax_id_extracted, 'expected': req.buyerTaxId}

        # 发票订单号
        platform_order = req.platformOrderNo
        remark_orders = extracted.get('remark_orders', [])
        order_passed = platform_order in remark_orders
        details['order_number'] = {'passed': order_passed, 'extracted': remark_orders, 'expected': platform_order}

        # 商品条码（仅非空）
        barcode_passed = bool(req.productBarcode)
        details['barcode'] = {'passed': barcode_passed, 'extracted': req.productBarcode, 'expected': '非空'}

        # SN码非空
        sn_passed = bool(req.snCode)
        details['sn_code'] = {'passed': sn_passed, 'extracted': req.snCode, 'expected': '非空'}

        # 商品品类白名单
        category_passed = False
        expected_category = ''
        if req.productCategory:
            # 电脑/笔记本电脑须人工审核（标记为不通过，但不属于退回意见自动填充范围）
            if '电脑' in req.productCategory or '笔记本电脑' in req.productCategory:
                category_passed = False
                expected_category = '电脑须人工审核'
            else:
                valid_categories = ['冰箱', '冰柜', '冰吧', '洗衣机', '电视', '激光电视', '空调', '中央空调',
                                    '热水器', '电热水器', '壁挂炉']  # 移除电脑相关
                category_passed = any(cat in req.productCategory for cat in valid_categories)
                expected_category = '属于白名单' if category_passed else '不属于白名单'
        else:
            expected_category = '未提供'
        details['product_category'] = {
            'passed': category_passed,
            'extracted': req.productCategory,
            'expected': expected_category
        }

        # 能耗等级
        energy_passed = req.energyGrade and (req.energyGrade == '1' or req.energyGrade == '一级' or req.energyGrade == '01')
        details['energy_grade'] = {'passed': energy_passed, 'extracted': req.energyGrade, 'expected': '一级'}

        # 发票订单金额（清理后比较）
        total_passed = True
        if req.totalAmount:
            extracted_total = extracted.get('remark_total')
            if extracted_total is None:
                total_passed = False
            else:
                cleaned_total = clean_number_str(extracted_total)
                if cleaned_total is None:
                    total_passed = False
                else:
                    try:
                        invoice_total = float(cleaned_total)
                        req_total = float(req.totalAmount)
                        total_passed = abs(invoice_total - req_total) < 0.01
                    except:
                        total_passed = False
        else:
            total_passed = True
        details['total_amount'] = {'passed': total_passed, 'extracted': extracted.get('remark_total'), 'expected': req.totalAmount}

        # ========== 新增：备案价校验 ==========
        price_check_passed = True
        price_limit = None
        if req.productBarcode and req.totalAmount:
            if req.productBarcode in price_map:
                price_limit = price_map[req.productBarcode]
                try:
                    order_price = float(req.totalAmount)
                    if order_price > price_limit:
                        price_check_passed = False
                except:
                    price_check_passed = False
            else:
                # 条码未找到，默认通过（不校验）
                price_check_passed = True
        else:
            price_check_passed = True
        details['price_limit'] = {
            'passed': price_check_passed,
            'extracted': price_limit,
            'expected': f'≤{price_limit}' if price_limit else '未找到备案价'
        }

        # 发票优惠金额
        discount_passed = True
        if req.discountAmount:
            extracted_discount = extracted.get('remark_discount')
            if extracted_discount is None:
                discount_passed = False
            else:
                cleaned_discount = clean_number_str(extracted_discount)
                if cleaned_discount is None:
                    discount_passed = False
                else:
                    try:
                        invoice_discount = float(cleaned_discount)
                        req_discount = float(req.discountAmount)
                        discount_passed = abs(invoice_discount - req_discount) < 0.01
                    except:
                        discount_passed = False
        else:
            discount_passed = True
        details['discount_amount'] = {'passed': discount_passed, 'extracted': extracted.get('remark_discount'), 'expected': req.discountAmount}

        # 发票实付金额
        paid_passed = True
        if req.amount:
            extracted_paid = extracted.get('remark_paid')
            if extracted_paid is None:
                paid_passed = False
            else:
                cleaned_paid = clean_number_str(extracted_paid)
                if cleaned_paid is None:
                    paid_passed = False
                else:
                    try:
                        invoice_paid = float(cleaned_paid)
                        req_paid = float(req.amount)
                        paid_passed = abs(invoice_paid - req_paid) < 0.01
                    except:
                        paid_passed = False
        else:
            paid_passed = True
        details['paid_amount'] = {'passed': paid_passed, 'extracted': extracted.get('remark_paid'), 'expected': req.amount}

        # 算术校验
        arithmetic_passed = True
        if req.totalAmount and req.discountAmount and req.amount:
            try:
                total = float(req.totalAmount)
                discount = float(req.discountAmount)
                paid = float(req.amount)
                arithmetic_passed = abs(total - discount - paid) < 0.01
            except:
                arithmetic_passed = False
        details['arithmetic_check'] = {'passed': arithmetic_passed, 'extracted': f"{req.totalAmount} - {req.discountAmount} = {req.amount}", 'expected': '相等'}

        # 签购单订单号（直接匹配或数字匹配）
        sign_order_passed = True
        if req.signPurchaseImageUrl or req.signPurchaseImageBase64:
            matched = False
            if sign_order and sign_order == req.platformOrderNo:
                matched = True
            if sign_external and sign_external == req.platformOrderNo:
                matched = True

            if not matched:
                def digits_only(s):
                    return re.sub(r'\D', '', s) if s else ''
                expected_digits = digits_only(req.platformOrderNo)
                if sign_order and digits_only(sign_order) == expected_digits:
                    matched = True
                if sign_external and digits_only(sign_external) == expected_digits:
                    matched = True

            sign_order_passed = matched
        else:
            sign_order_passed = True
        extracted_sign_order = None
        if sign_order and sign_external:
            extracted_sign_order = f"{sign_order} / {sign_external}"
        elif sign_order:
            extracted_sign_order = sign_order
        elif sign_external:
            extracted_sign_order = sign_external
        details['sign_order'] = {'passed': sign_order_passed, 'extracted': extracted_sign_order, 'expected': req.platformOrderNo}

        # 签购单原始金额
        sign_original_passed = True
        if req.signPurchaseImageUrl or req.signPurchaseImageBase64:
            if sign_original is None:
                sign_original_passed = False
            else:
                cleaned_original = clean_number_str(sign_original)
                if cleaned_original is None:
                    sign_original_passed = False
                else:
                    try:
                        if abs(float(cleaned_original) - float(req.totalAmount)) >= 0.01:
                            sign_original_passed = False
                    except:
                        sign_original_passed = False
        else:
            sign_original_passed = True
        details['sign_original'] = {'passed': sign_original_passed, 'extracted': sign_original, 'expected': req.totalAmount}

        # 签购单实付金额
        sign_paid_passed = True
        if req.signPurchaseImageUrl or req.signPurchaseImageBase64:
            if sign_paid is None:
                sign_paid_passed = False
            else:
                cleaned_paid = clean_number_str(sign_paid)
                if cleaned_paid is None:
                    sign_paid_passed = False
                else:
                    try:
                        if abs(float(cleaned_paid) - float(req.amount)) >= 0.01:
                            sign_paid_passed = False
                    except:
                        sign_paid_passed = False
        else:
            sign_paid_passed = True
        details['sign_paid'] = {'passed': sign_paid_passed, 'extracted': sign_paid, 'expected': req.amount}

        # 签购单优惠金额
        sign_discount_passed = True
        if req.signPurchaseImageUrl or req.signPurchaseImageBase64:
            if sign_discount is None:
                sign_discount_passed = False
            else:
                cleaned_discount = clean_number_str(sign_discount)
                if cleaned_discount is None:
                    sign_discount_passed = False
                else:
                    try:
                        if abs(float(cleaned_discount) - float(req.discountAmount)) >= 0.01:
                            sign_discount_passed = False
                    except:
                        sign_discount_passed = False
        else:
            sign_discount_passed = True
        details['sign_discount'] = {'passed': sign_discount_passed, 'extracted': sign_discount, 'expected': req.discountAmount}

        # SN码图片
        sn_image_passed = True
        if req.snImageUrl or req.snImageBase64:
            if sn_from_image is None:
                sn_image_passed = False
            elif sn_from_image != req.snCode:
                sn_image_passed = False
        else:
            sn_image_passed = True
        details['sn_image'] = {'passed': sn_image_passed, 'extracted': sn_from_image, 'expected': req.snCode}

        # SN码水印地址审核
        if sn_watermark_address:
            address_passed = '重庆' in sn_watermark_address
        else:
            address_passed = False
        details['sn_watermark_address'] = {
            'passed': address_passed,
            'extracted': sn_watermark_address,
            'expected': '包含重庆'
        }

        # 配送方式审核（线上平台自提不通过）
        delivery_method_passed = True
        delivery_method_expected = ''
        if req.platform == 'online':
            if req.deliveryMethod == '自提':
                delivery_method_passed = False
                delivery_method_expected = '自提须人工审核'
            else:
                delivery_method_expected = '非自提'
        else:
            # 线下平台无此字段，默认通过
            delivery_method_passed = True
            delivery_method_expected = '不适用'
        details['delivery_method'] = {
            'passed': delivery_method_passed,
            'extracted': req.deliveryMethod or '未获取',
            'expected': delivery_method_expected
        }
        # ---------- 物流审核结果（新增） ----------
        # 默认标记为通过（如果条件不满足则不在details中显示）
        if req.platform == 'online' and req.deliveryMethod != '自提' and req.logisticsImageUrl:
            # 物流单号核对
            tracking_passed = False
            extracted_tracking = logistics_fields.get('tracking_no')
            if req.trackingNo and extracted_tracking:
                if extracted_tracking == req.trackingNo:
                    tracking_passed = True
                else:
                    def clean_tracking(s):
                        return re.sub(r'[^A-Za-z0-9]', '', s)
                    if clean_tracking(extracted_tracking) == clean_tracking(req.trackingNo):
                        tracking_passed = True
            details['logistics_tracking'] = {'passed': tracking_passed, 'extracted': extracted_tracking, 'expected': req.trackingNo}

            # 收件人核对
            extracted_receiver = logistics_fields.get('receiver_name')
            receiver_passed = False
            # 若提取到收货人，则进行匹配；若未提取到，则默认通过
            if extracted_receiver:
                if req.buyerName:
                    pattern = '^' + re.escape(req.buyerName).replace('\\*', '.') + '$'
                    if re.match(pattern, extracted_receiver):
                        receiver_passed = True
                # 如果期望值为空（理论上不应发生），也可通过
                else:
                    receiver_passed = True
            else:
                # 未提取到收货人，默认通过（单项结果处会提示未提取到）
                receiver_passed = True
                print("⚠️ 未提取到收货人，自动标记为通过")

            details['logistics_receiver'] = {
                'passed': receiver_passed,
                'extracted': extracted_receiver,
                'expected': req.buyerName
            }

            # 地址在重庆
            address_passed = False
            extracted_address = logistics_fields.get('receiver_address')
            if extracted_address and '重庆' in extracted_address:
                address_passed = True
            details['logistics_address'] = {'passed': address_passed, 'extracted': extracted_address, 'expected': '包含重庆'}

            # 状态签收/完结
            status_passed = False
            extracted_status = logistics_fields.get('status', '')
            if extracted_status:
                finish_keywords = ['已送达', '已签收', '已妥投', '妥投', '送达', '签收', '已完成', '完结', '已派送', '派送完成', '派件完成']
                if any(kw in extracted_status for kw in finish_keywords) \
                    and not any(ek in extracted_status for ek in ['拒收', '退回', '未签收', '异常']):
                    status_passed = True
            details['logistics_status'] = {'passed': status_passed, 'extracted': extracted_status, 'expected': '签收/完结'}

        success = all(d['passed'] for d in details.values())

        extracted = make_json_serializable(extracted)
        details = make_json_serializable(details)

        return {
            "success": success,
            "details": details,
            "extracted": extracted,
            "debug": {
                "invoice_ocr": invoice_text[:1000] if invoice_text else "",
                "sign_ocr": sign_ocr[:1000] if 'sign_ocr' in locals() else "",
                "sn_ocr": sn_ocr[:1000] if 'sn_ocr' in locals() else "",
                "logistics_ocr": logistics_ocr[:1000] if 'logistics_ocr' in locals() else "",
            }
        }

    except Exception as e:
        print("❌ 发生异常，堆栈信息如下：")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# =============================================================================
# 健康检查
# =============================================================================
@app.get("/health")
def health():
    return {"status": "ok"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=5678)
